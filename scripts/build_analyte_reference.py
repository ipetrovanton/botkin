"""Сборка структурного справочника анализов из выгрузки ФСЛИ (xlsx).

Источник: «Справочник лабораторных тестов» ФСЛИ (OID 1.2.643.5.1.13.13.11.1080,
портал НСИ Минздрава). Один лист «Справочник», шапка — строка 2 (1-based), данные с строки 3.
Значимые колонки: LOINC, FULLNAME, ENGLISHNAME, SHORTNAME, SYNONYMS (через ';'),
UNIT, GROUP, TESTSTATUS, NMU.

Результат — registry.jsonl: по записи на тест с каноничным именем, краткой/английской формой,
синонимами, LOINC, кодом НМУ, единицей, группой и статусом. Позволяет в рантайме фаззи-коррекцию
названия по полному набору форм и заполнение LOINC/НМУ/ожидаемой единицы.

Запуск (сеть НЕ требуется):
    uv run python -m scripts.build_analyte_reference \\
        --src "Справочник лабораторных тестов.xlsx" \\
        --out src/botkin/reference/analytes/registry.jsonl
"""
from __future__ import annotations

import argparse
import json
from pathlib import Path

import openpyxl

STATUS_MAP = {"Актуальный": "active", "Новый": "new", "Устаревший": "deprecated"}
_MIN_NAME_LEN = 2
_HEADER_ROW = 2  # 1-based: строка с именами колонок


def normalize_key(name: str) -> str:
    """Ключ дедупликации/матчинга: lower, ё→е, схлопывание пробелов."""
    return " ".join(str(name).strip().lower().replace("ё", "е").split())


def _clean(value) -> str | None:
    if value is None:
        return None
    s = str(value).strip()
    return s or None


def _split_synonyms(value) -> list[str]:
    if value is None:
        return []
    return [s.strip() for s in str(value).split(";") if s.strip()]


def row_to_record(row: dict) -> dict | None:
    """Строка xlsx (dict по именам колонок) → запись реестра или None (если нет имени)."""
    full = _clean(row.get("FULLNAME"))
    if not full or len(full) < _MIN_NAME_LEN:
        return None
    loinc = _clean(row.get("LOINC"))
    if loinc == "0":
        loinc = None
    status_raw = _clean(row.get("TESTSTATUS")) or ""
    return {
        "name": full,
        "short": _clean(row.get("SHORTNAME")),
        "english": _clean(row.get("ENGLISHNAME")),
        "synonyms": _split_synonyms(row.get("SYNONYMS")),
        "loinc": loinc,
        "nmu": _clean(row.get("NMU")),
        "unit": _clean(row.get("UNIT")),
        "group": _clean(row.get("GROUP")),
        "specimen": _clean(row.get("SPECIMEN")),
        "status": STATUS_MAP.get(status_raw, status_raw.lower()),
    }


def build_registry(xlsx_path: Path) -> list[dict]:
    wb = openpyxl.load_workbook(xlsx_path, read_only=True)
    ws = wb["Справочник"] if "Справочник" in wb.sheetnames else wb.active
    if ws is None:
        wb.close()
        return []
    rows = ws.iter_rows(values_only=True)
    header = None
    for i, row in enumerate(rows, start=1):
        if i == _HEADER_ROW:
            header = [str(c).strip() if c is not None else "" for c in row]
            break
    if header is None:
        wb.close()
        return []

    seen: set[str] = set()
    out: list[dict] = []
    for row in rows:  # продолжаем с данных (после шапки)
        record = row_to_record(dict(zip(header, row)))
        if record is None:
            continue
        key = normalize_key(record["name"])
        if key in seen:
            continue
        seen.add(key)
        out.append(record)
    wb.close()
    return out


def write_registry(records: list[dict], out_path: Path, source_note: str) -> None:
    out_path.parent.mkdir(parents=True, exist_ok=True)
    with out_path.open("w", encoding="utf-8") as f:
        f.write(json.dumps({"_meta": source_note}, ensure_ascii=False) + "\n")
        for record in records:
            f.write(json.dumps(record, ensure_ascii=False) + "\n")


def main() -> None:  # pragma: no cover — ручной запуск с файлом выгрузки
    parser = argparse.ArgumentParser(description="Сборка справочника анализов botkin из ФСЛИ")
    parser.add_argument("--src", type=Path, required=True, help="xlsx-выгрузка ФСЛИ")
    parser.add_argument("--out", type=Path, required=True, help="Путь к registry.jsonl")
    args = parser.parse_args()

    records = build_registry(args.src)
    write_registry(records, args.out, source_note=f"ФСЛИ {args.src.name} ({len(records)})")
    print(f"Записано {len(records)} тестов в {args.out}")


if __name__ == "__main__":  # pragma: no cover
    main()
